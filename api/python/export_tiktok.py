import sys
import json

from openpyxl import load_workbook


# =========================
# ARGUMENTS
# =========================

template_path = sys.argv[1]
output_path = sys.argv[2]
json_data = sys.argv[3]

products = json.loads(json_data)


# =========================
# LOAD WORKBOOK
# =========================

wb = load_workbook(template_path)

ws = wb['Template']


# =========================
# FIND HEADER ROW
# =========================

header_row = None

for row in ws.iter_rows():

    values = [
        str(cell.value).strip()
        if cell.value is not None
        else ''
        for cell in row
    ]

    if (
        'SKU ID' in values
        and
        'Quantity' in values
    ):

        header_row = row[0].row
        break


if not header_row:

    raise Exception(
        'HEADER NOT FOUND'
    )


# =========================
# FIND COLUMN
# =========================

sku_col = None
qty_col = None

for cell in ws[header_row]:

    value = (
        str(cell.value)
        .strip()
        .lower()
    )

    if value == 'sku id':

        sku_col = cell.column

    if value == 'quantity':

        qty_col = cell.column


if not sku_col or not qty_col:

    raise Exception(
        'COLUMN NOT FOUND'
    )


# =========================
# UPDATE STOCK
# =========================

updated = 0

for row in ws.iter_rows(
    min_row=header_row + 1
):

    sku_id = row[
        sku_col - 1
    ].value

    if not sku_id:
        continue

    sku_id = (
        str(sku_id)
        .replace('.0', '')
        .replace(' ', '')
        .strip()
    )

    product = next(

        (
            item for item in products

            if
            str(
                item.get(
                    'tiktok_sku_id',
                    ''
                )
            )
            .replace('.0', '')
            .replace(' ', '')
            .strip()

            == sku_id
        ),

        None
    )

    if not product:
        continue

    # =========================
    # UPDATE QUANTITY
    # =========================

    row[
        qty_col - 1
    ].value = int(
        product.get(
            'stock',
            0
        )
    )

    updated += 1


# =========================
# SAVE
# =========================

wb.save(output_path)

print(
    json.dumps({

        'updated': updated,

        'outputPath':
            output_path
    })
)